import streamlit as st
import pandas as pd
import os
import time # For performance troubleshooting
from collections import defaultdict
import base64
from io import BytesIO
from datetime import datetime
import uuid
from PIL import Image, ImageOps
import plotly.express as px

# This assumes your firestore_manager.py file is present and correct
from firestore_manager import (
    integrate_with_streamlit_app,
    get_or_create_user_id,
    save_current_project_to_cloud,
    load_project_summaries_from_cloud,
    ensure_project_loaded,
)

# Initialize Firebase
firestore_manager = integrate_with_streamlit_app()


# --- SESSION STATE INITIALIZATION ---
if "projects" not in st.session_state:
    st.session_state.projects = {}  # full project data (loaded on demand)
if "project_summaries" not in st.session_state:
    st.session_state.project_summaries = [] # lightweight list for home screen

# Load project summaries on startup
load_project_summaries_from_cloud()

# Page state
if 'current_project' not in st.session_state:
    st.session_state.current_project = None
if 'page' not in st.session_state:
    st.session_state.page = 'projects'


# --- PAGE CONFIG AND STYLING ---
st.set_page_config(
    page_title="Interactive Product Grid",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Hide Streamlit style elements
hide_streamlit_style = """
<style>
    /* 1. Hide the "Hamburger" menu (top right) */
    #MainMenu {
        visibility: hidden;
        display: none;
    }

    /* 2. Hide the "Deploy" button (top right) */
    .stDeployButton {
        visibility: hidden;
        display: none;
    }

    /* 3. Hide the footer */
    footer {
        visibility: hidden;
        display: none;
    }

    /* 4. SAFETY: Ensure the Header Bar itself is visible 
       (This holds the sidebar toggle > arrow) */
    header[data-testid="stHeader"] {
        visibility: visible !important;
        background: transparent !important;
    }

    /* 5. Explicitly target the sidebar toggle arrow to ensure it's seen */
    [data-testid="stSidebarCollapsedControl"] {
        visibility: visible;
        display: block;
        color: inherit;
    }

    /* 6. Fix content padding so title isn't cut off */
    .block-container {
        padding-top: 3rem !important;
    }

    /* 7. Custom styles for the MultiSelect (Scrollable + Small Tags) */
    .stMultiSelect div[data-baseweb="select"] > div:first-child {
        max-height: 100px; 
        overflow-y: auto;
    }
    .stMultiSelect [data-baseweb="tag"] span {
        font-size: 12px !important; 
    }
    
    /* 8. NEW: Force Modal to be wider for editing */
    /* Target the dialog container to maximize its width */
    [data-testid="stModal"] > div > div {
        max-width: 90vw !important;
        width: 1400px !important;
    }
</style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# Custom CSS
st.markdown("""
<style>
    .main > div {
        padding-top: 2rem;
    }
    .stSelectbox > div > div > select {
        background-color: white;
    }
    .product-card {
        border: 1px solid #ddd;
        border-radius: 8px;
        padding: 9px;
        margin: 5px;
        background-color: white;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
        height: 100%;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
    }
    .product-card img {
        image-rendering: -webkit-optimize-contrast;
        image-rendering: crisp-edges;
        max-width: 100%;
        height: auto;
        margin-bottom: 10px;
    }
    .changed-attribute {
        color: #B22222;
        font-weight: bold;
    }
    .project-card {
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 20px;
        margin: 10px 0;
        background-color: white;
        box-shadow: 2px 2px 8px rgba(0,0,0,0.1);
        transition: transform 0.2s ease;
    }
    .project-card:hover {
        transform: translateY(-2px);
        box-shadow: 4px 4px 12px rgba(0,0,0,0.15);
    }
    .project-stats {
        color: #666;
        font-size: 0.9em;
    }
    .new-project-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 30px;
        border-radius: 15px;
        margin: 20px 0;
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

# --- IMAGE PROCESSING HELPERS ---
CARD_IMG_CSS_WIDTH = 200
CARD_IMG_CSS_HEIGHT = 220  # NEW: Fixed height for grid images
MODAL_IMG_CSS_WIDTH = 500 # UPDATED: Increased width for a zoomed-up look in modal
RETINA_FACTOR = 2


@st.cache_data(show_spinner=False)
def get_image_html_from_url(product_id: str, image_url: str, css_width: int):
    """Creates a simple <img> tag from a URL. Caches against the product_id."""
    if not image_url:
        # UPDATED: Use fixed height for placeholder
        return f'<div style="width: {css_width}px; height: {CARD_IMG_CSS_HEIGHT}px; display: flex; align-items: center; justify-content: center; background-color: #f0f2f6; border-radius: 8px;">📷 No image</div>'
    
    # UPDATED: Fixed height with object-fit: contain
    return f'<img src="{image_url}" style="width: auto; max-width: {css_width}px; height: {CARD_IMG_CSS_HEIGHT}px; object-fit: contain; display: block; margin-left: auto; margin-right: auto;" alt="Product Image">'

@st.cache_data(show_spinner=False)
def _encode_png_uri(im: Image.Image) -> str:
    """Return a PNG data URI for a PIL image."""
    b = BytesIO()
    im.save(b, format="PNG", optimize=True)
    return "data:image/png;base64," + base64.b64encode(b.getvalue()).decode("ascii")

@st.cache_data(show_spinner=False)
def _resize_lanczos(img: Image.Image, target_w: int) -> Image.Image:
    """High-quality downscale while preserving aspect ratio."""
    if img.width <= target_w:
        return img.copy()
    r = target_w / img.width
    new_h = max(1, int(round(img.height * r)))
    return img.resize((target_w, new_h), Image.Resampling.LANCZOS)

@st.cache_data(show_spinner=False)
def build_img_srcset(image_bytes: bytes, css_width: int) -> str:
    """Return an <img> HTML string with srcset (1x/2x) as PNG data URIs."""
    img = Image.open(BytesIO(image_bytes))
    img = ImageOps.exif_transpose(img)
    one_x = _resize_lanczos(img, css_width)
    two_x = _resize_lanczos(img, css_width * RETINA_FACTOR)
    uri_1x = _encode_png_uri(one_x)
    uri_2x = _encode_png_uri(two_x)
    
    # UPDATED: Fixed height with object-fit: contain
    return f"""
    <img
      src="{uri_1x}"
      srcset="{uri_1x} 1x, {uri_2x} {RETINA_FACTOR}x"
      style="width: auto; max-width: {css_width}px; height: {CARD_IMG_CSS_HEIGHT}px; object-fit: contain; display: block; margin-left: auto; margin-right: auto;"
      alt="Product Image"
    />
    """

@st.cache_data(show_spinner=False)
def get_cached_product_image_html(product_id: str, image_bytes: bytes, css_width: int):
    """Processes and caches the image HTML against the product_id."""
    if not image_bytes:
        # UPDATED: Use fixed height for placeholder
        return f'<div style="width: {css_width}px; height: {CARD_IMG_CSS_HEIGHT}px; display: flex; align-items: center; justify-content: center; background-color: #f0f2f6; border-radius: 8px;">📷 No image</div>'
    return build_img_srcset(image_bytes, css_width)


# --- CORE DATA FUNCTIONS ---
def create_new_project(name, description=""):
    """Create a new project with an empty data structure."""
    project_id = str(uuid.uuid4())
    project = {
        'id': project_id,
        'name': name,
        'description': description,
        'created_date': datetime.now().isoformat(),
        'last_modified': datetime.now().isoformat(),
        'products_data': [],
        'attributes': [],
        'distributions': [],
        'filter_options': {},
        'pending_changes': {},
        'uploaded_images': {},
        'excel_filename': None
    }
    st.session_state.projects[project_id] = project
    return project_id

def update_project_timestamp(project_id):
    """Update the last modified timestamp for a project."""
    if project_id in st.session_state.projects:
        st.session_state.projects[project_id]['last_modified'] = datetime.now().isoformat()

def sanitize_attr(attr):
    """Sanitize attribute names for use as keys."""
    return attr.replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '').replace('+', 'plus').replace(':', '')

def get_filter_options(products, attributes):
    """Generate filter options from products data."""
    opts = defaultdict(set)
    for p in products:
        for attr in attributes:
            opts[attr].add(p["attributes"][attr])
    return {k: sorted(v) for k, v in opts.items()}

def find_image_for_product(product_id, uploaded_images):
    """Find matching image for a product ID."""
    for filename, file_data in uploaded_images.items():
        name_part = os.path.splitext(filename)[0].lower()
        if name_part == str(product_id).lower():
            return file_data
    return None

def create_image_lookup(uploaded_images: dict) -> dict:
    """
    Creates a dictionary for instant image lookups.
    Key: product_id (lowercase string), Value: image_bytes.
    This is much faster than looping.
    """
    return {
        os.path.splitext(filename)[0].lower(): file_data
        for filename, file_data in uploaded_images.items()
    }

def load_and_parse_excel(uploaded_file, image_url_mappings):
    """
    More robustly parse an uploaded Excel file and return structured data,
    mapping products to their IMAGE URLs instead of raw bytes.
    """
    if uploaded_file is None:
        return [], [], [], {}

    # --- OPTIMIZATION: Create the URL lookup dictionary ONCE ---
    # The keys are product IDs, and the values are the image URLs.
    url_lookup = {}
    for product_id, meta in image_url_mappings.items():
        if isinstance(meta, dict) and meta.get("public_url"):
            url_lookup[product_id.lower()] = meta["public_url"]
        elif isinstance(meta, str): # Legacy support
            url_lookup[product_id.lower()] = meta
    
    try:
        df = pd.read_excel(uploaded_file)
        df.columns = [str(c).strip() for c in df.columns]

        id_col = next((c for c in df.columns if 'product id' in c.lower()), None)
        desc_col = next((c for c in df.columns if 'description' in c.lower()), None)
        price_col = next((c for c in df.columns if 'price' in c.lower()), None)

        if not id_col or not desc_col:
            st.error("❌ Critical Error: Could not find 'Product ID' and 'Description' columns in the Excel file.")
            st.stop()
        
        attributes = [h for h in df.columns if h.startswith("ATT")]
        distributions = [h for h in df.columns if h.startswith("DIST")]
        
        products = []
        for idx, row in df.iterrows():
            
            # --- ROBUST ID HANDLING ---
            raw_id = row[id_col]
            # If Excel gave us a float like 123.0, convert to integer 123 first
            if isinstance(raw_id, float) and raw_id.is_integer():
                product_id = str(int(raw_id)).strip()
            else:
                product_id = str(raw_id).strip()
            
            # Double check for string ".0" suffix just in case
            if product_id.endswith(".0"):
                product_id = product_id[:-2]
            # --------------------------

            description = str(row[desc_col]).strip()
            
            try:
                price_val = float(row[price_col]) if price_col and not pd.isna(row[price_col]) else 0.0
                price_str = f"{price_val:.2f}"
            except (ValueError, TypeError):
                price_str = "0.00"
            
            # --- OPTIMIZATION: Use the fast dictionary lookup to get the URL ---
            image_url = url_lookup.get(product_id.lower())
            
            attr_data = {a: str(row[a]).strip() for a in attributes}
            dist_data = {d: "X" in str(row[d]).upper() for d in distributions}
            
            products.append({
                "original_index": idx, "product_id": product_id, 
                "image_data": None, # We no longer use raw bytes here
                "image_url": image_url, # We now use the URL
                "description": description, "original_description": description,
                "price": price_str, "original_price": price_str,
                "attributes": attr_data, "original_attributes": attr_data.copy(),
                "distribution": dist_data
            })
        
        filter_options = get_filter_options(products, attributes)
        return products, attributes, distributions, filter_options

    except Exception as e:
        st.error(f"❌ Failed to parse the Excel file. Please check its format. Error: {e}")
        return [], [], [], {}
        
        
def apply_filters(products, attribute_filters, distribution_filters, pending_changes=None, show_pending_only=False):
    """Apply filters to products and return filtered list."""
    if not products:
        return []

    filtered_products = products

    # --- NEW: Filter by Pending Changes ---
    if show_pending_only and pending_changes:
        # Normalize keys to strings to ensure we catch both int/str formats
        pending_keys = set(str(k) for k in pending_changes.keys())
        filtered_products = [
            p for p in filtered_products 
            if str(p["original_index"]) in pending_keys
        ]
    # --------------------------------------
    
    for attr, selected_values in attribute_filters.items():
        if selected_values and 'All' not in selected_values:
            filtered_products = [
                p for p in filtered_products if p["attributes"].get(attr, "") in selected_values
            ]

    if distribution_filters and 'All' not in distribution_filters:
        filtered_products = [
            p for p in filtered_products if any(
                p["distribution"].get(orig_dist, False)
                for dist in distribution_filters
                for orig_dist in p["distribution"] if orig_dist.replace('DIST ', '') == dist
            )
        ]
        
    return filtered_products

    
def create_download_excel(project):
    """Create Excel file, highlighting ONLY the most recently applied changes."""
    if not project['products_data']:
        return None
    
    attributes = project['attributes']
    distributions = project['distributions']
    headers = ["Product ID", "Description"] + attributes + distributions + ["Price"]
    
    data = []
    for product in project['products_data']:
        row = [product["product_id"], product["description"]]
        row.extend(product["attributes"].get(attr, "") for attr in attributes)
        row.extend("X" if product["distribution"].get(dist, False) else "" for dist in distributions)
        try:
            p_val = float(product["price"]) if product["price"] else 0.0
        except:
            p_val = 0.0
        row.append(p_val)
        data.append(row)
    
    df = pd.DataFrame(data, columns=headers)
    
    from openpyxl.styles import PatternFill
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Products')
        workbook = writer.book
        ws = writer.sheets['Products']
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Retrieve the history of the last batch of applied changes
        last_applied = project.get('last_applied_changes', {})
        
        for i, product in enumerate(project['products_data']):
            excel_row = i + 2
            
            # Get changes for this specific product ID (handle int vs string keys)
            idx = product['original_index']
            changes = last_applied.get(idx) or last_applied.get(str(idx)) or {}
            
            # If this product wasn't in the last batch, skip it
            if not changes:
                continue

            # --- Check Description (Column 2) ---
            if 'description' in changes:
                ws.cell(row=excel_row, column=2).fill = yellow_fill
            
            # --- Check Attributes (Columns 3 to 3+len(attrs)) ---
            for j, attr in enumerate(attributes):
                if attr in changes:
                    col_idx = 3 + j
                    ws.cell(row=excel_row, column=col_idx).fill = yellow_fill
            
            # --- Check Price (Last Column) ---
            if 'price' in changes:
                price_col = len(headers)
                ws.cell(row=excel_row, column=price_col).fill = yellow_fill

    return output.getvalue()


# --- UI DISPLAY FUNCTIONS ---
def display_product_card(product, project, visible_attributes):
    """Display a single product card."""
    with st.container(border=True):
        
        image_html = get_image_html_from_url(
            product_id=product["product_id"], 
            image_url=product.get("image_url"), 
            css_width=CARD_IMG_CSS_WIDTH
        )
        st.markdown(image_html, unsafe_allow_html=True)
        
        # --- Check Pending Changes for Red Highlighting ---
        idx = product["original_index"]
        pending = project.get('pending_changes', {})
        product_changes = pending.get(idx) or pending.get(str(idx)) or {}
        # --------------------------------------------------

        card_content = ""
        
        # --- NEW: Always Show Product ID ---
        # This sits above the description in gray text
        card_content += f'<div style="font-size: 11px; color: #888; margin-bottom: 2px;">ID: {product["product_id"]}</div>'
        # -----------------------------------

        if "Description" in visible_attributes:
            desc_class = "changed-attribute" if "description" in product_changes else ""
            card_content += f'<p class="{desc_class}" style="margin-bottom: 4px;"><strong>{product["description"]}</strong></p>'
        
        if "Price" in visible_attributes and product.get("price"):
            price_class = "changed-attribute" if "price" in product_changes else ""
            card_content += f'<p class="{price_class}" style="margin-bottom: 8px;">Price: ${product["price"]}</p>'
        
        # Attributes Loop
        for attr in project.get('attributes', []):
            if attr in visible_attributes:
                current_val = product["attributes"].get(attr, "N/A")
                is_changed = attr in product_changes
                
                style = 'color: #B22222; font-weight: bold;' if is_changed else ''
                clean_attr = attr.replace('ATT ', '')
                
                card_content += f'<div style="font-size: 12px; line-height: 1.4; {style}"><strong>{clean_attr}:</strong> {current_val}</div>'
        
        st.markdown(card_content, unsafe_allow_html=True)

        if not st.session_state.get("client_mode", False):
            if st.button(f"Edit", key=f"edit_{product['original_index']}_{project['id']}", use_container_width=True):
                st.session_state.editing_product = product
                st.rerun()

def show_edit_modal(product, project):
    @st.dialog(f"Edit Product: {product['product_id']}")
    def edit_product_dialog():
        
        # --- FIXED LAYOUT FOR IMAGE AND ATTRIBUTES ---
        # RATIO: [3, 2] for 60% Image and 40% Form. Modal width is forced wider by CSS.
        col_img, col_form = st.columns([3, 2]) 
        
        with col_img:
            st.subheader("Product Image")
            if product.get("image_url"):
                # MODAL_IMG_CSS_WIDTH is 500 (zoomed)
                st.image(
                    product["image_url"], 
                    caption=f"ID: {product['product_id']}", 
                    width=MODAL_IMG_CSS_WIDTH
                )
            else:
                st.info("No image available.")
        
        with col_form:
            # --- Product Details Section ---
            st.subheader("Product Details")
            # Description and Price in a horizontal pair
            col1, col2 = st.columns(2)
            new_description = col1.text_input("Description", value=product["description"])
            new_price = col2.text_input("Price", value=product["price"])
            
            st.markdown("---") # Visual separator to break up the flow
            
            # --- Attributes Section (This is the critical area) ---
            st.subheader("Attributes")
            
            # Attributes in a two-column sub-grid inside the right column
            # By nesting the attributes inside a fixed-width column (col_form), 
            # and using the 2 sub-columns for *pairs* of attributes, 
            # we ensure the list flows vertically and horizontally without bleeding.
            attr_cols = st.columns(2)
            new_attributes = {}
            for i, attr in enumerate(project['attributes']):
                with attr_cols[i % 2]:
                    current_val = product["attributes"][attr]
                    options = project['filter_options'].get(attr, [])
                    if current_val not in options:
                        options.insert(0, current_val)
                    
                    clean_attr = attr.replace('ATT ', '')
                    index = options.index(current_val) if current_val in options else 0
                    
                    selected_option = st.selectbox(
                        f"{clean_attr}", options + ["[Custom Value]"], index=index,
                        key=f"modal_attr_{attr}_{product['original_index']}"
                    )
                    
                    if selected_option == "[Custom Value]":
                        new_attributes[attr] = st.text_input(
                            f"Custom {clean_attr}", value=current_val, 
                            key=f"modal_custom_{attr}_{product['original_index']}"
                        )
                    else:
                        new_attributes[attr] = selected_option
        # --- END NEW LAYOUT ---

        st.markdown("---")
        _, save_col, cancel_col, _ = st.columns([1, 1, 1, 1])
        if save_col.button("💾 Save Changes", type="primary", use_container_width=True):
            idx = product["original_index"]
            if idx not in project['pending_changes']:
                project['pending_changes'][idx] = {}

            if new_description != product["description"]:
                project['pending_changes'][idx]["description"] = new_description
                product["description"] = new_description
            
            if new_price != product["price"]:
                project['pending_changes'][idx]["price"] = new_price
                product["price"] = new_price
            
            for attr, new_val in new_attributes.items():
                if new_val != product["attributes"][attr]:
                    project['pending_changes'][idx][attr] = new_val
                    product["attributes"][attr] = new_val
                    if new_val not in project['filter_options'].get(attr, []):
                        project['filter_options'].setdefault(attr, []).append(new_val)
                        project['filter_options'][attr].sort()
            
            update_project_timestamp(project['id'])
            st.success("✅ Changes saved!")
            del st.session_state.editing_product
            st.rerun()
        
        if cancel_col.button("❌ Cancel", use_container_width=True):
            del st.session_state.editing_product
            st.rerun()

    edit_product_dialog()


# --- PAGE VIEW FUNCTIONS ---
def show_projects_page():
    """Display the main projects page."""
    st.title("📁 Interactive Product Grid")

    with st.container():
        st.markdown('<div class="new-project-section"><h2>🚀 Start a New Project</h2><p>Create a new product grid project with your Excel data and images</p></div>', unsafe_allow_html=True)
        if st.button("➕ Create New Project", type="primary", use_container_width=True):
            st.session_state.page = 'create_project'
            st.rerun()

    summaries = sorted(st.session_state.get("project_summaries", []), key=lambda s: s.get('last_modified', ''), reverse=True)

    if summaries:
        st.header("📋 Your Projects")
        for s in summaries:
            with st.container(border=True):
                col1, col2, col3 = st.columns([5, 1, 1])
                with col1:
                    st.subheader(s.get("name", "Untitled"))
                    if s.get("description"): st.write(s.get("description"))
                    try: lm_disp = datetime.fromisoformat(s['last_modified']).strftime("%Y-%m-%d %H:%M")
                    except: lm_disp = "—"
                    st.markdown(f"""
                    <div class="project-stats">
                        📦 {s.get("num_products", 0)} products | 
                        🏷️ {s.get("num_attributes", 0)} attributes | 
                        ✏️ {s.get("num_pending_changes", 0)} pending | 
                        🕒 Modified: {lm_disp}
                    </div>
                    """, unsafe_allow_html=True)
                
                if col2.button("Open", key=f"open_{s['id']}", type="primary", use_container_width=True):
                    if ensure_project_loaded(s['id']):
                        st.query_params["project"] = s['id']
                        st.session_state.current_project = s['id']
                        st.session_state.page = 'grid'
                        st.rerun()
                        return
                    else: st.error("Could not load project.")

                if col3.button("🗑️", key=f"delete_{s['id']}", use_container_width=True):
                    mgr = st.session_state.get("firestore_manager")
                    if mgr and mgr.delete_project(s['id']):
                        st.session_state.project_summaries = [p for p in st.session_state.project_summaries if p["id"] != s['id']]
                        if s['id'] in st.session_state.projects: del st.session_state.projects[s['id']]
                        if st.session_state.current_project == s['id']: st.session_state.current_project = None
                        st.rerun()

def show_create_project_page():
    """Display the create new project page."""
    st.title("🆕 Create New Project")
    if st.button("← Back to Projects"):
        st.session_state.page = 'projects'
        st.rerun()
    
    with st.form(key="new_project_form"):
        project_name = st.text_input("Project Name *", placeholder="e.g., Q1 2024 Product Launch")
        
        # --- REMOVED: Description Input ---
        
        col1, col2 = st.columns(2)
        uploaded_excel = col1.file_uploader("Upload Product Excel Grid *", type=['xlsx', 'xls'])
        uploaded_images = col2.file_uploader("Upload Product Images", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True)
        
        # --- Match Preview Logic ---
        if uploaded_excel and uploaded_images:
            try:
                df_preview = pd.read_excel(uploaded_excel)
                df_preview.columns = [str(c).strip() for c in df_preview.columns]
                id_col = next((c for c in df_preview.columns if 'product id' in c.lower()), None)
                
                if id_col:
                    excel_ids = set()
                    for x in df_preview[id_col].dropna():
                        if isinstance(x, float) and x.is_integer():
                            excel_ids.add(str(int(x)).strip().lower())
                        else:
                            clean_id = str(x).strip().lower()
                            if clean_id.endswith(".0"): clean_id = clean_id[:-2]
                            excel_ids.add(clean_id)
                    
                    image_names = set(os.path.splitext(img.name)[0].strip().lower() for img in uploaded_images)
                    
                    matches = excel_ids.intersection(image_names)
                    match_count = len(matches)
                    
                    if match_count == 0:
                        st.error(f"⚠️ **0 matches found!** Check your filenames. (Excel has {len(excel_ids)} IDs, Uploaded {len(image_names)} images)")
                        st.write("Example Excel ID:", list(excel_ids)[0] if excel_ids else "None")
                        st.write("Example Filename:", list(image_names)[0] if image_names else "None")
                    else:
                        st.success(f"✅ **{match_count}** images matched out of {len(uploaded_images)} uploaded.")
            except Exception:
                pass 
        
        submitted = st.form_submit_button("🚀 Create Project", type="primary", use_container_width=True)
        
        if submitted:
            if not project_name or not uploaded_excel:
                st.warning("Please provide a project name and an Excel file.")
            else:
                with st.spinner("Creating project and uploading assets..."):
                    # 1. Parse the Excel file. 
                    excel_bytes = uploaded_excel.getvalue()
                    products, attrs, dists, filters = load_and_parse_excel(BytesIO(excel_bytes), {})
                    
                    # 2. Manually match uploaded images
                    if uploaded_images:
                        product_lookup = {p['product_id'].lower().strip(): p for p in products}
                        for img in uploaded
